"""
Importa Simulador Jera Endowment.xlsm para fixtures JSON.

Mapeamento 17 classes granulares -> 8 buckets:
  BR_CASH_FI    = Liquidez CDI + RF BR Credito Pos + RF BR Pre + RF BR Inflacao
  BR_EQUITIES   = Renda Variavel BR
  BR_ALTS       = Retorno Absoluto BR + Private Equity BR
  BR_RE         = Real Estate BR
  INT_CASH_FI   = Cash Equivalent + RF Intl Pre + RF Intl Inflacao + RF Intl Credito
  INT_EQUITIES  = Renda Variavel Intl
  INT_ALTS      = Retorno Absoluto Intl + Private Equity Intl + Commodities
  INT_RE        = Real Estate Intl

Emite:
  prisma/fixtures/asset-classes.json  (8 buckets)
  prisma/fixtures/profiles.json       (12 perfis, Pesos Perfis → 8 buckets)
  prisma/fixtures/historical-returns.json (serie mensal por bucket,
                                           media equal-weight dos constituintes)
"""
from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path

import openpyxl

XLSM = Path(
    r"C:\Users\fnobre\OneDrive - JERA CAPITAL GESTAO DE RECURSOS LTDA\Cópia de Simulador Jera Endowment.xlsm"
)
OUT_DIR = Path(__file__).resolve().parent.parent / "prisma" / "fixtures"
OUT_DIR.mkdir(parents=True, exist_ok=True)

# Exact column names from Retornos sheet (normalized: accents replaced)
# Order matches Excel columns 2-18.
RAW_CLASSES_ORDER = [
    ("RF BR CDI", 2),
    ("Cash Equivalent", 3),
    ("RF BR Credito Pos-Fixado", 4),
    ("RF BR Pre-Fixado", 5),
    ("RF BR Inflacao", 6),
    ("RF Intl. Pre-Fixado", 7),
    ("RF Intl. Inflacao", 8),
    ("RF Intl. Credito", 9),
    ("Retorno Absoluto BR", 10),
    ("Retorno Absoluto Intl.", 11),
    ("Renda Variavel BR", 12),
    ("Renda Variavel Intl.", 13),
    ("Private Equity BR", 14),
    ("Private Equity Intl.", 15),
    ("Real Estate BR", 16),
    ("Real Estate Intl.", 17),
    ("Commodities", 18),
]

# 8 buckets + constituintes
BUCKET_DEFS = {
    "br_cash_fi": {
        "label": "BR Cash + Renda Fixa",
        "region": "brasil",
        "category": "cash_fi",
        "constituents": [
            "RF BR CDI",
            "RF BR Credito Pos-Fixado",
            "RF BR Pre-Fixado",
            "RF BR Inflacao",
        ],
    },
    "br_equities": {
        "label": "BR Renda Variavel",
        "region": "brasil",
        "category": "equities",
        "constituents": ["Renda Variavel BR"],
    },
    "br_alts": {
        "label": "BR Alternativos",
        "region": "brasil",
        "category": "alternatives",
        "constituents": ["Retorno Absoluto BR", "Private Equity BR"],
    },
    "br_re": {
        "label": "BR Real Estate",
        "region": "brasil",
        "category": "real_estate",
        "constituents": ["Real Estate BR"],
    },
    "int_cash_fi": {
        "label": "INT Cash + Renda Fixa",
        "region": "internacional",
        "category": "cash_fi",
        "constituents": [
            "Cash Equivalent",
            "RF Intl. Pre-Fixado",
            "RF Intl. Inflacao",
            "RF Intl. Credito",
        ],
    },
    "int_equities": {
        "label": "INT Renda Variavel",
        "region": "internacional",
        "category": "equities",
        "constituents": ["Renda Variavel Intl."],
    },
    "int_alts": {
        "label": "INT Alternativos",
        "region": "internacional",
        "category": "alternatives",
        "constituents": [
            "Retorno Absoluto Intl.",
            "Private Equity Intl.",
            "Commodities",
        ],
    },
    "int_re": {
        "label": "INT Real Estate",
        "region": "internacional",
        "category": "real_estate",
        "constituents": ["Real Estate Intl."],
    },
}


def read_returns() -> list[dict]:
    """Retorna lista de {date: ISO, by_class: {classe: retorno}}."""
    wb = openpyxl.load_workbook(XLSM, data_only=True, read_only=True)
    ws = wb["Retornos"]
    out: list[dict] = []

    # Header na linha 3; dados a partir da linha 4
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row or row[0] is None:
            continue
        date = row[0]
        if not isinstance(date, datetime):
            continue
        by_class: dict[str, float] = {}
        for name, col_idx in RAW_CLASSES_ORDER:
            val = row[col_idx - 1]
            if isinstance(val, (int, float)):
                by_class[name] = float(val)
        if by_class:
            out.append({"date": date.strftime("%Y-%m-%d"), "by_class": by_class})

    wb.close()
    return out


def aggregate_to_buckets(monthly: list[dict]) -> list[dict]:
    """Media equal-weight dos constituintes de cada bucket por mes."""
    out: list[dict] = []
    for entry in monthly:
        by_bucket: dict[str, float] = {}
        for bucket_id, defn in BUCKET_DEFS.items():
            parts = [
                entry["by_class"][c]
                for c in defn["constituents"]
                if c in entry["by_class"]
            ]
            if parts:
                by_bucket[bucket_id] = sum(parts) / len(parts)
        out.append({"date": entry["date"], "returns": by_bucket})
    return out


def read_profiles() -> list[dict]:
    """12 perfis da aba 'Pesos Perfis' -> 8 buckets (soma dos constituintes)."""
    wb = openpyxl.load_workbook(XLSM, data_only=True, read_only=True)
    ws = wb["Pesos Perfis"]

    # Row 2: regioes (Brasil / Internacional / Br & Intl)
    # Row 3: "Classe" + nomes dos 12 perfis
    # Row 4+: classe | peso_perfil_1 | peso_perfil_2 | ...
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    regions_row = rows[1]
    levels_row = rows[2]

    # Cols B..M (index 1..12) contem os 12 perfis
    profile_cols = []
    for i in range(1, 13):
        region = regions_row[i] if i < len(regions_row) else None
        level = levels_row[i] if i < len(levels_row) else None
        if not level or not region:
            continue
        profile_cols.append(
            {
                "col_index": i,
                "region": str(region).lower(),
                "level": str(level).lower(),
                "id": f"{str(region).lower().replace(' & ', '_').replace(' ', '_')}_{str(level).lower()}",
            }
        )

    # Constroi mapa nome_do_excel -> classe_normalizada (para bater com RAW_CLASSES_ORDER)
    excel_to_norm = {
        "Liquidez CDI": "RF BR CDI",
        "Cash Equivalent": "Cash Equivalent",
        "RF BR Credito Pos-Fixado": "RF BR Credito Pos-Fixado",
        "RF BR Pre-Fixado": "RF BR Pre-Fixado",
        "RF BR Inflacao": "RF BR Inflacao",
        "RF Intl. Pre-Fixado": "RF Intl. Pre-Fixado",
        "RF Intl. Inflacao": "RF Intl. Inflacao",
        "RF Intl. Credito": "RF Intl. Credito",
        "Retorno Absoluto BR": "Retorno Absoluto BR",
        "Retorno Absoluto Intl.": "Retorno Absoluto Intl.",
        "Renda Variavel BR": "Renda Variavel BR",
        "Renda Variavel Intl.": "Renda Variavel Intl.",
        "Private Equity BR": "Private Equity BR",
        "Private Equity Intl.": "Private Equity Intl.",
        "Real Estate BR": "Real Estate BR",
        "Real Estate Intl.": "Real Estate Intl.",
        "Commodities": "Commodities",
    }

    # Pesos brutos por perfil: dict[profile_id][class_name] = weight
    raw_weights: dict[str, dict[str, float]] = {
        p["id"]: {} for p in profile_cols
    }

    for row in rows[3:]:
        if not row or not row[0]:
            continue
        class_name_excel = str(row[0]).strip()
        # Normalize accents (the row[0] came with accents; compare simplified)
        import unicodedata
        norm = "".join(
            c for c in unicodedata.normalize("NFD", class_name_excel) if unicodedata.category(c) != "Mn"
        )
        if norm not in excel_to_norm:
            continue
        class_key = excel_to_norm[norm]
        for p in profile_cols:
            w = row[p["col_index"]]
            if isinstance(w, (int, float)):
                raw_weights[p["id"]][class_key] = float(w)

    # Fold 17-class weights -> 8 buckets (sum of constituents)
    profiles_out = []
    for p in profile_cols:
        bucket_weights = {}
        for bucket_id, defn in BUCKET_DEFS.items():
            bucket_weights[bucket_id] = sum(
                raw_weights[p["id"]].get(c, 0) for c in defn["constituents"]
            )
        total = sum(bucket_weights.values())
        profiles_out.append(
            {
                "id": p["id"],
                "region": p["region"],
                "level": p["level"],
                "weights": bucket_weights,
                "total": round(total, 6),
            }
        )
    return profiles_out


def main():
    print(f"Lendo {XLSM.name}...")

    # Asset classes
    classes = [
        {
            "id": bid,
            "label": d["label"],
            "region": d["region"],
            "category": d["category"],
            "constituents": d["constituents"],
        }
        for bid, d in BUCKET_DEFS.items()
    ]
    (OUT_DIR / "asset-classes.json").write_text(
        json.dumps(classes, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(f"  ✓ asset-classes.json ({len(classes)} buckets)")

    # Historical returns
    print("  Lendo Retornos...")
    monthly = read_returns()
    bucket_series = aggregate_to_buckets(monthly)
    (OUT_DIR / "historical-returns.json").write_text(
        json.dumps(bucket_series, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(f"  ✓ historical-returns.json ({len(bucket_series)} meses)")

    # Profiles
    print("  Lendo Pesos Perfis...")
    profiles = read_profiles()
    (OUT_DIR / "profiles.json").write_text(
        json.dumps(profiles, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(f"  ✓ profiles.json ({len(profiles)} perfis)")

    # Stats rapidas de sanidade
    print("\n--- Sanidade ---")
    print(f"Periodo: {bucket_series[0]['date']} -> {bucket_series[-1]['date']}")
    print("Pesos do perfil 'br_&_intl_agressivo':")
    target = next((p for p in profiles if p["id"].endswith("agressivo") and "intl" in p["id"]), None)
    if target:
        for k, v in target["weights"].items():
            print(f"  {k}: {v:.4f}")
        print(f"  total = {target['total']}")


if __name__ == "__main__":
    main()
